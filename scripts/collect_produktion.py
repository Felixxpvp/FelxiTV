"""
Daten: Produktions-Sammler

Liest scripts/produktion.xlsx (Met, Gemüse, Obst, Jagd & Fleisch)
und speichert die Summen als Tagesstand in der Datenbank.

Erzeugt die Tabellen: produktion_met, produktion_gemuese, produktion_obst, produktion_jagd
"""

import sys
from datetime import datetime, timezone
from pathlib import Path

WORKSPACE_ROOT = Path(__file__).resolve().parent.parent
EXCEL_PATH = WORKSPACE_ROOT / "scripts" / "produktion.xlsx"


def _sicherer_wert(v):
    """Gibt den Wert als Zahl zurück, 0 wenn None oder leer."""
    if v is None:
        return 0
    try:
        return float(v)
    except (ValueError, TypeError):
        return 0


def collect():
    """Produktionsdaten aus der Excel-Datei lesen."""
    if not EXCEL_PATH.exists():
        return {"source": "produktion", "status": "skipped",
                "reason": f"Datei nicht gefunden: {EXCEL_PATH}"}

    try:
        import openpyxl
        wb = openpyxl.load_workbook(str(EXCEL_PATH), data_only=True)

        def tab_summen(sheet_name, idx_menge, idx_verkauft):
            """Summen aus einem Tab holen."""
            if sheet_name not in wb.sheetnames:
                return 0, 0
            ws = wb[sheet_name]
            menge = sum(_sicherer_wert(r[idx_menge]) for r in ws.iter_rows(min_row=2, values_only=True)
                        if any(v is not None for v in r))
            verkauft = sum(_sicherer_wert(r[idx_verkauft]) for r in ws.iter_rows(min_row=2, values_only=True)
                           if any(v is not None for v in r))
            return menge, verkauft

        met_prod, met_verk = tab_summen("Met", 2, 3)
        gem_prod, gem_verk = tab_summen("Gemüse", 2, 3)
        obst_prod, obst_verk = tab_summen("Obst", 2, 3)
        jagd_prod, jagd_verk = tab_summen("Jagd & Fleisch", 2, 3)

        return {
            "source": "produktion",
            "status": "success",
            "data": {
                "met_produziert": int(met_prod),
                "met_verkauft": int(met_verk),
                "gemuese_produziert": round(gem_prod, 1),
                "gemuese_verkauft": round(gem_verk, 1),
                "obst_produziert": round(obst_prod, 1),
                "obst_verkauft": round(obst_verk, 1),
                "jagd_produziert": round(jagd_prod, 1),
                "jagd_verkauft": round(jagd_verk, 1),
            }
        }
    except Exception as e:
        return {"source": "produktion", "status": "error", "reason": str(e)}


def write(conn, result, date):
    """Produktionsdaten in die Datenbank schreiben."""
    conn.execute("""
        CREATE TABLE IF NOT EXISTS produktion_stand (
            date TEXT PRIMARY KEY,
            met_produziert INTEGER DEFAULT 0,
            met_verkauft INTEGER DEFAULT 0,
            gemuese_produziert REAL DEFAULT 0,
            gemuese_verkauft REAL DEFAULT 0,
            obst_produziert REAL DEFAULT 0,
            obst_verkauft REAL DEFAULT 0,
            jagd_produziert REAL DEFAULT 0,
            jagd_verkauft REAL DEFAULT 0,
            collected_at TEXT
        )
    """)

    d = result["data"]
    conn.execute("""
        INSERT OR REPLACE INTO produktion_stand
        (date, met_produziert, met_verkauft, gemuese_produziert, gemuese_verkauft,
         obst_produziert, obst_verkauft, jagd_produziert, jagd_verkauft, collected_at)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
    """, (date, d["met_produziert"], d["met_verkauft"],
          d["gemuese_produziert"], d["gemuese_verkauft"],
          d["obst_produziert"], d["obst_verkauft"],
          d["jagd_produziert"], d["jagd_verkauft"],
          datetime.now(timezone.utc).isoformat()))
    conn.commit()
    return 1


if __name__ == "__main__":
    result = collect()
    print(f"Status: {result['status']}")
    if result["status"] == "success":
        d = result["data"]
        print(f"  Met:    {d['met_produziert']} prod. / {d['met_verkauft']} verk.")
        print(f"  Gemüse: {d['gemuese_produziert']} kg prod. / {d['gemuese_verkauft']} kg verk.")
        print(f"  Obst:   {d['obst_produziert']} kg prod. / {d['obst_verkauft']} kg verk.")
        print(f"  Jagd:   {d['jagd_produziert']} kg prod. / {d['jagd_verkauft']} kg verk.")
    else:
        print(f"  {result.get('reason')}")
