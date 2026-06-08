"""
Daten: Finanz-Sammler

Liest scripts/finanzen.xlsx (Einnahmen und Ausgaben)
und speichert Tagesstände nach Sparte in der Datenbank.

Erzeugt die Tabelle: finanzen_stand
"""

from datetime import datetime, timezone
from pathlib import Path

WORKSPACE_ROOT = Path(__file__).resolve().parent.parent
EXCEL_PATH = WORKSPACE_ROOT / "scripts" / "finanzen.xlsx"


def _sicherer_wert(v):
    if v is None:
        return 0.0
    try:
        return float(v)
    except (ValueError, TypeError):
        return 0.0


def _tab_nach_sparte(wb, sheet_name):
    """Beträge aus einem Tab gruppiert nach Sparte summieren."""
    if sheet_name not in wb.sheetnames:
        return {}
    ws = wb[sheet_name]
    summen = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not any(v is not None for v in row):
            continue
        sparte = str(row[2]).strip() if row[2] else "Sonstiges"
        betrag = _sicherer_wert(row[3])
        summen[sparte] = summen.get(sparte, 0.0) + betrag
    return summen


def collect():
    """Finanzdaten aus der Excel-Datei lesen."""
    if not EXCEL_PATH.exists():
        return {"source": "finanzen", "status": "skipped",
                "reason": f"Datei nicht gefunden: {EXCEL_PATH}"}

    try:
        import openpyxl
        wb = openpyxl.load_workbook(str(EXCEL_PATH), data_only=True)

        einnahmen = _tab_nach_sparte(wb, "Einnahmen")
        ausgaben = _tab_nach_sparte(wb, "Ausgaben")

        gesamt_ein = sum(einnahmen.values())
        gesamt_aus = sum(ausgaben.values())

        return {
            "source": "finanzen",
            "status": "success",
            "data": {
                "einnahmen_gesamt": round(gesamt_ein, 2),
                "ausgaben_gesamt": round(gesamt_aus, 2),
                "ergebnis": round(gesamt_ein - gesamt_aus, 2),
                "einnahmen_nach_sparte": einnahmen,
                "ausgaben_nach_sparte": ausgaben,
            }
        }
    except Exception as e:
        return {"source": "finanzen", "status": "error", "reason": str(e)}


def write(conn, result, date):
    """Finanzdaten in die Datenbank schreiben."""
    conn.execute("""
        CREATE TABLE IF NOT EXISTS finanzen_stand (
            date TEXT PRIMARY KEY,
            einnahmen_gesamt REAL DEFAULT 0,
            ausgaben_gesamt REAL DEFAULT 0,
            ergebnis REAL DEFAULT 0,
            einnahmen_json TEXT,
            ausgaben_json TEXT,
            collected_at TEXT
        )
    """)

    import json
    d = result["data"]
    conn.execute("""
        INSERT OR REPLACE INTO finanzen_stand
        (date, einnahmen_gesamt, ausgaben_gesamt, ergebnis,
         einnahmen_json, ausgaben_json, collected_at)
        VALUES (?, ?, ?, ?, ?, ?, ?)
    """, (date, d["einnahmen_gesamt"], d["ausgaben_gesamt"], d["ergebnis"],
          json.dumps(d["einnahmen_nach_sparte"], ensure_ascii=False),
          json.dumps(d["ausgaben_nach_sparte"], ensure_ascii=False),
          datetime.now(timezone.utc).isoformat()))
    conn.commit()
    return 1


if __name__ == "__main__":
    result = collect()
    print(f"Status: {result['status']}")
    if result["status"] == "success":
        d = result["data"]
        print(f"  Einnahmen gesamt: {d['einnahmen_gesamt']:.2f} €")
        print(f"  Ausgaben gesamt:  {d['ausgaben_gesamt']:.2f} €")
        print(f"  Ergebnis:         {d['ergebnis']:.2f} €")
        print("  Einnahmen nach Sparte:", d["einnahmen_nach_sparte"])
        print("  Ausgaben nach Sparte: ", d["ausgaben_nach_sparte"])
    else:
        print(f"  {result.get('reason')}")
